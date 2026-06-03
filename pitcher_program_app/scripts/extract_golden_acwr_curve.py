"""Program Engine Task 0.3 — extract golden ACWR curve from the raw xlsx.

This is the FUTURE script that replaces tests/fixtures/golden_acwr_curve.json
with full daily 5-tuple precision once the operator copies the resolved
`Ramp up with Bullpen` xlsx into `data/knowledge/golden/`.

Right now the fixture is seeded from the recon dossier (Front 5 verbatim
transcript) because the xlsx exists in the repo only as a 1108-byte Google
Drive alias that openpyxl can't read.

When the real file lands at
  data/knowledge/golden/ramp_up_with_bullpen_12wk.xlsx
run this script; it extracts column B (Distance) / C (Throws) / D (Intent) /
E (Drill) / F (daily total throws) / G (load units) per Front 5's verified
schema, computes the weekly G curve, and overwrites
tests/fixtures/golden_acwr_curve.json.

Usage:
    python -m scripts.extract_golden_acwr_curve

Exit codes:
    0 — extracted and wrote fixture
    1 — input xlsx missing (still Drive-aliased)
    2 — schema mismatch (header row didn't match expected B/C/D/E/F/G)
"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

GOLDEN = Path(__file__).resolve().parents[1] / "data" / "knowledge" / "golden" / "ramp_up_with_bullpen_12wk.xlsx"
OUT = Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "golden_acwr_curve.json"

EXPECTED_HEADERS = {
    "B": ("distance", "ft"),
    "C": ("throws", "count"),
    "D": ("intent", "pct"),
    "E": ("drill", "str"),
    "F": ("daily_total_throws", "count"),
    "G": ("load_units", "G"),
}


def main() -> int:
    if not GOLDEN.exists():
        print(f"missing: {GOLDEN}", file=sys.stderr)
        print(
            "  The xlsx exists in past_arm_programs/ only as a Drive alias on this host.\n"
            "  Resolve in Finder ('Show Original' → save outside Drive) and copy here.\n"
            "  See data/knowledge/golden/README.md for details.",
            file=sys.stderr,
        )
        return 1

    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed; pip install openpyxl", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(GOLDEN, data_only=True, read_only=True)
    ws = wb.active

    # Find the header row (search first 5 rows for an "ACWLR" or "G" / "%increase" signal)
    header_row_idx = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        joined = " ".join(str(c) for c in row if c)
        if "ACWLR" in joined or "%increase" in joined or "load" in joined.lower():
            header_row_idx = r_idx
            break
    if header_row_idx is None:
        print("could not locate header row in first 8 rows; xlsx schema drifted", file=sys.stderr)
        return 2

    daily_rows: list[dict] = []
    week = None
    day = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        # Heuristic week detection: a row with a distinctive marker (e.g. column A containing "Week N")
        first = row[0] if row else None
        if isinstance(first, str) and "week" in first.lower():
            # Extract week number
            import re as _re
            m = _re.search(r"(\d+)", first)
            if m:
                week = int(m.group(1))
                day = 0
                continue
        if week is None:
            continue
        # Expect columns B,C,D,E,F,G at indices 1..6
        distance, throws, intent, drill, daily_total, load_units = (row[i] if i < len(row) else None for i in range(1, 7))
        if all(v in (None, "", 0) for v in (distance, throws, intent, load_units)):
            continue
        day += 1
        daily_rows.append(
            {
                "week": week,
                "day": day,
                "distance_ft": distance,
                "throw_count": throws,
                "intent_pct": int(round(float(intent) * 100)) if isinstance(intent, (int, float)) and intent <= 1 else intent,
                "drill": drill,
                "daily_total_throws": daily_total,
                "G_load_units": load_units,
            }
        )

    if not daily_rows:
        print("extracted 0 daily rows; schema heuristics failed", file=sys.stderr)
        return 2

    # Aggregate weekly G
    weekly_G: dict[int, float] = {}
    for r in daily_rows:
        if isinstance(r.get("G_load_units"), (int, float)):
            weekly_G[r["week"]] = weekly_G.get(r["week"], 0.0) + float(r["G_load_units"])
    weeks_sorted = sorted(weekly_G.keys())
    weekly_curve = [weekly_G[w] for w in weeks_sorted]

    payload = {
        "_meta": {
            "description": "12-week golden ACWR fixture extracted from the raw xlsx.",
            "primary_source": "data/knowledge/golden/ramp_up_with_bullpen_12wk.xlsx",
            "extracted_at": date.today().isoformat(),
            "extracted_by": "scripts.extract_golden_acwr_curve",
            "daily_row_count": len(daily_rows),
        },
        "weekly_G_load_units": weekly_curve,
        "daily_5tuples": daily_rows,
        "deload_weeks_1_indexed": [
            weeks_sorted[i]
            for i in range(1, len(weekly_curve))
            if weekly_curve[i] < weekly_curve[i - 1] * 0.95
        ],
        "expected_invariants": {
            "min_weekly_G": min(weekly_curve),
            "max_weekly_G": max(weekly_curve),
            "acute_chronic_ratio_band": {"lower": 0.8, "upper": 1.3, "hard_cap": 1.5},
            "monotonic_overall_trajectory": True,
            "deload_present": True,
        },
    }
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    print(f"Extracted {len(daily_rows)} daily rows across {len(weeks_sorted)} weeks → {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

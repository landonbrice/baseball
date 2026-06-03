"""Program Engine Task 0.1 — audit alias coverage of the golden xlsx assets.

Walks every cell in data/knowledge/golden/*.xlsx and classifies each free-text
exercise-name-looking string against the live `exercises.aliases` index. Emits
a report of (a) what resolved, (b) what didn't.

Unresolved names need a decision per row:
  - "alias to add" — the string names a known exercise; UPDATE exercises
    SET aliases = aliases || '[<new alias>]' WHERE id = <existing_id>.
  - "new exercise" — the string names something the library doesn't have;
    add a row via the existing seed pipeline.
  - "noise" — header / instruction / not an exercise name at all; ignore.

The script does NOT make changes. It prints a report. Operator decides.

Usage:
    python -m scripts.audit_golden_alias_coverage [--csv out.csv]

Exit codes:
    0 — full coverage (no unresolved names)
    1 — unresolved names found (the expected first-run state)
    2 — error loading goldens or building the index
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable

import openpyxl

from bot.services import exercise_alias

# Golden files we mine. Order is presentation-only.
GOLDEN_DIR = Path(__file__).resolve().parents[1] / "data" / "knowledge" / "golden"
GOLDEN_FILES = [
    "the_program.xlsx",
    "maintenance_plan.xlsx",
    "periodized_lifting.xlsx",
]

# Cells that obviously aren't exercise names. Pre-filter so the unresolved
# bucket doesn't drown in headers / instructions / cueing prose.
_NOISE_PATTERNS = [
    re.compile(r"^\s*$"),
    # Pure prescriptions: "3X8", "2X10 ea", "X10", "10 sec hold", "down and back"
    re.compile(r"^\d+([./xX-]\d+)*\s*(reps?|sec|min|each|set|ea)?\s*$", re.IGNORECASE),
    re.compile(r"^x\d+", re.IGNORECASE),
    re.compile(r"^down(\s+and\s+back)?$", re.IGNORECASE),
    # Section / block / phase / day labels
    re.compile(r"^\s*(day\s*\d+|block\s*\d+|week\s*\d+|phase\s*\d+|set\s*/\s*rep|comments?|warmup|rest|focus|comment)\b", re.IGNORECASE),
    re.compile(r"^(superset column|exercise\s*$|reps?\s*$|weight)\b", re.IGNORECASE),
    re.compile(r"^(superset|ss(\s+with)?|circuit|conditioning|recovery|maintenance plan|hypertrophy|strength|strength-power|posterior chain|anterior chain|push/pull|push pull|day [a-z])\s*$", re.IGNORECASE),
    re.compile(r"^phase\s+\d+:", re.IGNORECASE),
    re.compile(r"^a\d|^b\d|^c\d|^d\d|^ss$", re.IGNORECASE),
    # Cueing prose: contains a comma/period/semicolon and is long, OR contains conjunction words
    re.compile(r".+[,;:].+", re.IGNORECASE),  # any cell with a comma/semicolon/colon
    re.compile(r"\b(focus on|do not|continue with|spend more time|increase volume|lower volume|stay tall|sit back|use j bands?|use uninvolved|away leg|each side|use water ball|or modified|or assisted|other options?)\b", re.IGNORECASE),
    # Common notes / verbose descriptors
    re.compile(r"^.{60,}$"),  # cells > 60 chars are almost certainly instructions, not exercise names
    # Goal / dynamic warmup / post throw stretch — block names, not exercise names
    re.compile(r"^(post[\- ]?throw\s+stretch|dynamic\s+warmup|pre\s+bullpen\s+plyo|long\s+toss\s+to\s+comfort|pitch|goal:.*|abbreviated)$", re.IGNORECASE),
    re.compile(r"^\s*\d+\s*laps?\b", re.IGNORECASE),
    re.compile(r"^\d+\s+yrd\s*$", re.IGNORECASE),
    re.compile(r"^begins\s+at\b", re.IGNORECASE),
]


def _is_noise(s: str) -> bool:
    s = s.strip()
    return any(p.match(s) for p in _NOISE_PATTERNS)


def _looks_like_exercise_name(s: str) -> bool:
    if not isinstance(s, str):
        return False
    s = s.strip()
    if not s or len(s) < 3 or len(s) > 80:
        return False
    if _is_noise(s):
        return False
    # Exercise names almost always have letters and rarely consist of only digits/punct
    if not re.search(r"[A-Za-z]", s):
        return False
    return True


def _harvest_strings(path: Path) -> list[tuple[str, str, str]]:
    """Yield (sheet_name, cell_addr, raw_string) for every plausibly-name-shaped cell."""
    rows: list[tuple[str, str, str]] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        print(f"  ! failed to load {path}: {exc}", file=sys.stderr)
        return rows
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and _looks_like_exercise_name(val):
                    rows.append((sn, cell.coordinate, val.strip()))
    return rows


def _bucket(names: Iterable[tuple[str, str, str, str]]) -> tuple[dict, dict]:
    """Partition harvested names into resolved (with ex_id) and unresolved."""
    resolved: dict[str, list[tuple[str, str, str]]] = {}  # ex_id → [(file, sheet, raw), ...]
    unresolved: dict[str, list[tuple[str, str, str]]] = {}  # normalized → [(file, sheet, raw), ...]
    for file, sheet, addr, raw in names:
        ex_id = exercise_alias.try_resolve_alias(raw)
        if ex_id is not None:
            resolved.setdefault(ex_id, []).append((file, sheet, raw))
        else:
            try:
                key = exercise_alias._normalize(raw)
            except exercise_alias.UnknownExerciseAlias:
                continue
            unresolved.setdefault(key, []).append((file, sheet, raw))
    return resolved, unresolved


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--csv", help="If set, write the full report as CSV to this path.")
    p.add_argument(
        "--snapshot",
        default=str(Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "exercises_snapshot.json"),
        help="JSON snapshot of the exercises table to use instead of Supabase. Default: tests/fixtures/exercises_snapshot.json",
    )
    p.add_argument("--live", action="store_true",
                   help="Use live Supabase instead of the snapshot. Requires SUPABASE_URL + SUPABASE_SERVICE_KEY in env.")
    args = p.parse_args()

    if not args.live:
        try:
            exercise_alias.load_from_snapshot(args.snapshot)
            print(f"Loaded alias index from snapshot: {args.snapshot}")
        except FileNotFoundError:
            print(f"  ! snapshot missing: {args.snapshot}", file=sys.stderr)
            print(f"    Run `python -m scripts.dump_exercises_snapshot` to regenerate", file=sys.stderr)
            return 2

    print(f"Auditing alias coverage across {GOLDEN_DIR}/")
    all_names: list[tuple[str, str, str, str]] = []  # (file, sheet, addr, raw)
    for fname in GOLDEN_FILES:
        path = GOLDEN_DIR / fname
        if not path.exists():
            print(f"  ! missing: {path}", file=sys.stderr)
            continue
        harvested = _harvest_strings(path)
        print(f"  · {fname}: {len(harvested)} name-shaped cells")
        for sheet, addr, raw in harvested:
            all_names.append((fname, sheet, addr, raw))

    if not all_names:
        print("No name-shaped cells found. Goldens missing or filters too aggressive.", file=sys.stderr)
        return 2

    try:
        resolved, unresolved = _bucket(all_names)
    except Exception as exc:
        print(f"audit failed: {exc}", file=sys.stderr)
        return 2

    print()
    print(f"=== COVERAGE ===")
    print(f"  resolved   : {sum(len(v) for v in resolved.values())} cell-references → {len(resolved)} unique exercises")
    print(f"  unresolved : {sum(len(v) for v in unresolved.values())} cell-references → {len(unresolved)} unique normalized names")
    print()

    if unresolved:
        print(f"=== UNRESOLVED NAMES (decide per row: alias-to-add | new-exercise | noise) ===")
        # Sort by frequency desc — the most-cited unknowns are likeliest exercises
        for norm_key, occurrences in sorted(unresolved.items(), key=lambda kv: -len(kv[1])):
            raws_seen = sorted({r for _, _, r in occurrences})
            files_seen = sorted({f for f, _, _ in occurrences})
            print(f"  ({len(occurrences):2d}x) {raws_seen[0]!r}")
            if len(raws_seen) > 1:
                print(f"        also seen as: {raws_seen[1:]}")
            print(f"        in: {files_seen}")
        print()

    if args.csv:
        with open(args.csv, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["status", "ex_id_or_norm_key", "file", "sheet", "raw_text"])
            for ex_id, occurrences in sorted(resolved.items()):
                for file, sheet, raw in occurrences:
                    w.writerow(["resolved", ex_id, file, sheet, raw])
            for norm_key, occurrences in sorted(unresolved.items()):
                for file, sheet, raw in occurrences:
                    w.writerow(["unresolved", norm_key, file, sheet, raw])
        print(f"CSV written: {args.csv}")

    return 0 if not unresolved else 1


if __name__ == "__main__":
    sys.exit(main())

"""Unified research resolver — single door for all surfaces.

Replaces the split routing logic in knowledge_retrieval.py. All surfaces
(plan_gen, coach_chat, morning, daily_plan_why) call resolve_research()
to get research context for a pitcher.
"""

import logging
import os
import re
from dataclasses import dataclass, field
from typing import Literal, Optional

import yaml

from bot.config import KNOWLEDGE_DIR
from bot.services.vocabulary import (
    INJURY_AREAS,
    get_all_trigger_keywords,
    get_research_triggers_for_injury,
    get_research_triggers_for_mod,
)

logger = logging.getLogger(__name__)

RESEARCH_DIR = os.path.join(KNOWLEDGE_DIR, "research")

ALL_CONTEXTS = ["plan_gen", "coach_chat", "morning", "daily_plan_why", "program_gen"]


@dataclass
class DocRef:
    id: str
    title: str
    summary: str
    priority: str
    contexts: list[str] = field(default_factory=list)


@dataclass
class ResearchPayload:
    combined_text: str
    loaded_docs: list[DocRef]
    trigger_reason: str


_index_cache: dict[str, tuple[dict, str]] = {}


def _parse_frontmatter(text: str) -> dict:
    match = re.match(r"^---\s*\n(.*?)\n---", text, re.DOTALL)
    if not match:
        return {
            "id": "", "title": "", "keywords": [], "type": "",
            "applies_to": [], "triggers": [], "phase": "any",
            "priority": "standard", "contexts": list(ALL_CONTEXTS),
            "summary": "",
        }

    try:
        parsed = yaml.safe_load(match.group(1)) or {}
    except yaml.YAMLError:
        parsed = {}

    return {
        "id": parsed.get("id", ""),
        "title": parsed.get("title", ""),
        "keywords": parsed.get("keywords", []),
        "type": parsed.get("type", ""),
        "applies_to": parsed.get("applies_to", []),
        "triggers": parsed.get("triggers", []),
        "phase": parsed.get("phase", "any"),
        "priority": parsed.get("priority", "standard"),
        "contexts": parsed.get("contexts", list(ALL_CONTEXTS)),
        "summary": parsed.get("summary", ""),
    }


def _strip_frontmatter(text: str) -> str:
    return re.sub(r"^---\s*\n.*?\n---\s*\n?", "", text, count=1, flags=re.DOTALL)


def _load_index() -> dict[str, tuple[dict, str]]:
    global _index_cache
    if _index_cache:
        return _index_cache

    if not os.path.exists(RESEARCH_DIR):
        logger.warning("Research directory not found: %s", RESEARCH_DIR)
        return {}

    for filename in os.listdir(RESEARCH_DIR):
        if not filename.endswith(".md") or filename == "INDEX.md":
            continue
        filepath = os.path.join(RESEARCH_DIR, filename)
        try:
            with open(filepath, "r") as f:
                content = f.read()
            fm = _parse_frontmatter(content)
            doc_id = fm["id"] or filename.replace(".md", "")
            body = _strip_frontmatter(content)
            _index_cache[doc_id] = (fm, body)

            if not fm["id"]:
                logger.warning("Research doc %s missing 'id' in frontmatter", filename)
        except Exception as e:
            logger.warning("Error reading research file %s: %s", filename, e)

    logger.info("Research index: %d docs loaded", len(_index_cache))
    return _index_cache


def clear_cache():
    global _index_cache
    _index_cache = {}


def get_citations_for_ids(doc_ids: list[str]) -> list[dict]:
    """Resolve a list of research doc IDs to citation cards.

    Returns [{id, title, summary}] for each id that exists in the index.
    Unknown ids are silently dropped (template seeds may reference docs that
    have been renamed or removed). Used by the Program Builder /finalize
    endpoint to surface "why this program" references in the preview state.
    """
    if not doc_ids:
        return []
    index = _load_index()
    out = []
    for doc_id in doc_ids:
        entry = index.get(doc_id)
        if not entry:
            continue
        fm, _ = entry
        out.append({
            "id": doc_id,
            "title": fm.get("title", ""),
            "summary": fm.get("summary", ""),
        })
    return out


def should_fire_research(
    pitcher_profile: dict,
    triage_result: dict | None = None,
    user_message: str | None = None,
) -> tuple[bool, str]:
    if triage_result:
        flag = triage_result.get("flag_level", "green")
        if flag in ("yellow", "red", "modified_green"):
            return True, f"flag_level={flag}"

        mods = triage_result.get("modifications", [])
        if mods:
            first_mod = mods[0] if isinstance(mods[0], str) else str(mods[0])
            return True, f"recent_mod:{first_mod}"

    active_mods = (pitcher_profile.get("active_flags") or {}).get("active_modifications", [])
    if active_mods:
        return True, f"recent_mod:{active_mods[0]}"

    if user_message:
        msg_lower = user_message.lower()
        trigger_keywords = get_all_trigger_keywords()
        for kw in trigger_keywords:
            if kw in msg_lower:
                return True, f"keyword:{kw}"

    return False, ""


def resolve_research(
    pitcher_profile: dict,
    context: Literal["plan_gen", "coach_chat", "morning", "daily_plan_why"],
    triage_result: dict | None = None,
    user_message: str | None = None,
    max_chars: int = 12000,
) -> ResearchPayload:
    index = _load_index()
    loaded: dict[str, tuple[DocRef, str]] = {}

    injury_areas = set()
    for injury in pitcher_profile.get("injury_history", []):
        area = injury.get("area", "").lower()
        if area:
            injury_areas.add(area)

    injury_triggers = set()
    for area in injury_areas:
        injury_triggers.update(get_research_triggers_for_injury(area))

    mod_triggers = set()
    if triage_result:
        for mod in triage_result.get("modifications", []):
            mod_key = mod if isinstance(mod, str) else str(mod)
            mod_triggers.update(get_research_triggers_for_mod(mod_key))

    active_mods = (pitcher_profile.get("active_flags") or {}).get("active_modifications", [])
    for mod in active_mods:
        mod_triggers.update(get_research_triggers_for_mod(mod))

    trigger_reasons = []

    # Step 1: Critical docs for this context where applies_to matches
    for doc_id, (fm, content) in index.items():
        if context not in fm.get("contexts", ALL_CONTEXTS):
            continue
        if fm["priority"] != "critical":
            continue
        applies = fm.get("applies_to", [])
        if "any" in applies or injury_areas.intersection(applies):
            ref = DocRef(
                id=doc_id, title=fm["title"], summary=fm["summary"],
                priority=fm["priority"], contexts=fm.get("contexts", ALL_CONTEXTS),
            )
            loaded[doc_id] = (ref, content)
            trigger_reasons.append(f"critical:{doc_id}")

    # Step 2: Docs whose triggers intersect triage modifications
    all_triggers = injury_triggers | mod_triggers
    if all_triggers:
        for doc_id, (fm, content) in index.items():
            if doc_id in loaded:
                continue
            if context not in fm.get("contexts", ALL_CONTEXTS):
                continue
            doc_triggers = set(fm.get("triggers", []))
            if doc_triggers.intersection(all_triggers):
                ref = DocRef(
                    id=doc_id, title=fm["title"], summary=fm["summary"],
                    priority=fm["priority"], contexts=fm.get("contexts", ALL_CONTEXTS),
                )
                loaded[doc_id] = (ref, content)
                trigger_reasons.append(f"trigger_match:{doc_id}")

    # Step 3: User message keyword match (coach_chat only)
    if user_message and context == "coach_chat":
        msg_lower = user_message.lower()
        for doc_id, (fm, content) in index.items():
            if doc_id in loaded:
                continue
            if context not in fm.get("contexts", ALL_CONTEXTS):
                continue
            doc_triggers = set(fm.get("triggers", []))
            if any(t in msg_lower for t in doc_triggers):
                ref = DocRef(
                    id=doc_id, title=fm["title"], summary=fm["summary"],
                    priority=fm["priority"], contexts=fm.get("contexts", ALL_CONTEXTS),
                )
                loaded[doc_id] = (ref, content)
                trigger_reasons.append(f"keyword_match:{doc_id}")

    # Step 4: Standard docs matching injury_areas (fill remaining budget)
    for doc_id, (fm, content) in index.items():
        if doc_id in loaded:
            continue
        if context not in fm.get("contexts", ALL_CONTEXTS):
            continue
        if fm["priority"] != "standard":
            continue
        applies = fm.get("applies_to", [])
        if injury_areas.intersection(applies):
            ref = DocRef(
                id=doc_id, title=fm["title"], summary=fm["summary"],
                priority=fm["priority"], contexts=fm.get("contexts", ALL_CONTEXTS),
            )
            loaded[doc_id] = (ref, content)
            trigger_reasons.append(f"standard:{doc_id}")

    # Build combined text with budget
    combined_parts = []
    total_chars = 0
    final_docs = []
    for doc_id, (ref, content) in loaded.items():
        if total_chars + len(content) > max_chars:
            if ref.priority == "critical":
                remaining = max_chars - total_chars
                combined_parts.append(content[:remaining])
                final_docs.append(ref)
            break
        combined_parts.append(content)
        total_chars += len(content)
        final_docs.append(ref)

    combined = "\n\n---\n\n".join(combined_parts)
    reason = "; ".join(trigger_reasons) if trigger_reasons else "baseline"

    _log_research_load(
        pitcher_profile.get("pitcher_id") or pitcher_profile.get("id", ""),
        context, reason,
        [d.id for d in final_docs],
        len(combined),
    )

    if final_docs:
        logger.info(
            "Research resolved [%s]: %d docs (%s) — %d chars",
            context, len(final_docs),
            ", ".join(d.id for d in final_docs), len(combined),
        )

    return ResearchPayload(
        combined_text=combined,
        loaded_docs=final_docs,
        trigger_reason=reason,
    )


def _log_research_load(
    pitcher_id: str, context: str, trigger_reason: str,
    doc_ids: list[str], total_chars: int, degraded: bool = False,
):
    if not pitcher_id:
        return
    try:
        from bot.services.db import get_client
        get_client().table("research_load_log").insert({
            "pitcher_id": pitcher_id,
            "context": context,
            "trigger_reason": trigger_reason,
            "loaded_doc_ids": doc_ids,
            "total_chars": total_chars,
            "degraded": degraded,
        }).execute()
    except Exception as e:
        logger.debug("research_load_log insert failed (non-blocking): %s", e)


# ─────────────────────────────────────────────────────────────────────────────
# Program Engine — program_gen context (Task 1.3)
# ─────────────────────────────────────────────────────────────────────────────


GOLDEN_DIR = os.path.join(KNOWLEDGE_DIR, "golden")


def _load_golden_exemplars() -> list[dict]:
    """Return structured-name representations of every xlsx in data/knowledge/golden/.

    Lazy import openpyxl + a thin sheet→list-of-dict converter so importing
    this module doesn't pay the cost on every server boot. Returns an empty
    list (NOT raises) if the directory or any individual file is unreadable —
    a missing exemplar shouldn't break generation, only reduce richness.
    """
    if not os.path.isdir(GOLDEN_DIR):
        return []
    try:
        import openpyxl  # type: ignore
    except ImportError:
        logger.warning("openpyxl not available; golden exemplars cannot be loaded")
        return []

    out: list[dict] = []
    for filename in sorted(os.listdir(GOLDEN_DIR)):
        if not filename.endswith(".xlsx"):
            continue
        path = os.path.join(GOLDEN_DIR, filename)
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            logger.warning("could not load golden exemplar %s: %s", filename, e)
            continue
        sheets: list[dict] = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows: list[list] = []
            for r in ws.iter_rows(values_only=True):
                # Skip fully-empty rows to keep payloads tight
                if any(c is not None and str(c).strip() for c in r):
                    rows.append([str(c) if c is not None else "" for c in r])
            sheets.append({"sheet_name": sn, "rows": rows})
        out.append({"name": filename, "sheets": sheets})
    return out


def _knowledge_version(docs: list[tuple[str, str]], templates: list[dict], exemplars: list[dict]) -> str:
    """Deterministic SHA-1 hex over the resolver-assembled knowledge pack.

    `PitcherProgram.knowledge_version` persists this so the "edit a doc →
    next generation differs" proof falls out of the hash changing. Order is
    normalized by sorting on doc_id; templates are hashed by template_id +
    a stable JSON-dumps of content + tunable + mods + research_doc_ids;
    exemplars by name + row-count + the joined first-row of each sheet.
    """
    import hashlib
    import json as _json
    h = hashlib.sha1()
    for doc_id, text in sorted(docs, key=lambda x: x[0]):
        h.update(doc_id.encode("utf-8"))
        h.update(b"\x00")
        h.update(text.encode("utf-8"))
        h.update(b"\x01")
    for tpl in sorted(templates, key=lambda t: t.get("block_template_id") or ""):
        h.update((tpl.get("block_template_id") or "").encode("utf-8"))
        for k in ("content", "tunable_parameters_schema", "modification_rules_json", "research_doc_ids", "goal_tags"):
            h.update(_json.dumps(tpl.get(k), sort_keys=True, default=str).encode("utf-8"))
        h.update(b"\x02")
    for ex in sorted(exemplars, key=lambda x: x.get("name") or ""):
        h.update((ex.get("name") or "").encode("utf-8"))
        for sheet in ex.get("sheets") or []:
            h.update((sheet.get("sheet_name") or "").encode("utf-8"))
            h.update(str(len(sheet.get("rows") or [])).encode("utf-8"))
        h.update(b"\x03")
    return h.hexdigest()


def resolve_for_program_gen(
    pitcher_profile: dict,
    pitcher_context: str,
    goal_spec: dict,
    max_chars: int = 16000,
) -> dict:
    """Assemble goal-relevant knowledge for Program Engine v1 authoring (Task 1.3).

    Args:
        pitcher_profile: profile dict (must include `injury_history` if filtering by injury).
        pitcher_context: the per-pitcher context.md text (unused at v1 except
            in logging — reserved for the resolver to learn pitcher voice).
        goal_spec: at minimum `{"tags": ["velocity"], ...}`. Other keys are
            passed through to the LLM later.
        max_chars: total chars budget for `combined` text. Default 16k = ~3×
            the legacy resolver because program authoring is one-shot, not
            per-day.

    Returns:
        {
            "docs":       [(doc_id: str, text: str), ...],
            "templates":  [{block_template_id, content, tunable_parameters_schema,
                           modification_rules_json, research_doc_ids, goal_tags,
                           compatible_phases, duration_range_weeks}, ...],
            "exemplars":  [{name, sheets: [{sheet_name, rows: [[...]]}]}, ...],
            "knowledge_version": "<sha1 hex>",
            "combined": "<doc_id>\\n...\\n---\\n<doc_id>\\n..."  (budget-bounded),
            "loaded_doc_ids": [doc_id, ...],
        }

    Selection rule for program_gen context:
        (a) every doc whose frontmatter `triggers` intersects `goal_spec.tags`
        (b) every doc with `priority: critical` (always-on safety/physiology baseline)
        (c) every doc whose `applies_to` matches the pitcher's injury history
        (d) every doc whose `contexts` includes `program_gen` (operator opt-in)
        Then keep adding by priority (critical→standard→reference) until budget.

    Templates:
        Every block_library row whose `goal_tags` intersects `goal_spec.tags`.

    Exemplars:
        All loadable xlsx in data/knowledge/golden/.
    """
    tags = set(t.lower() for t in (goal_spec.get("tags") or []))
    pitcher_id = pitcher_profile.get("pitcher_id") or pitcher_profile.get("id") or ""

    # Pitcher injury areas (from injury_history)
    injury_areas: set[str] = set()
    for inj in pitcher_profile.get("injury_history") or []:
        area = (inj or {}).get("area") or (inj or {}).get("injury_area")
        if area:
            injury_areas.add(str(area).lower())

    index = _load_index()
    selected: list[tuple[str, dict, str]] = []  # (doc_id, fm, body)
    seen_ids: set[str] = set()

    def _maybe_add(doc_id: str, fm: dict, body: str):
        if doc_id in seen_ids:
            return
        seen_ids.add(doc_id)
        selected.append((doc_id, fm, body))

    for doc_id, (fm, body) in index.items():
        contexts = fm.get("contexts", ALL_CONTEXTS)
        # Hard filter: doc must opt-in to program_gen
        if "program_gen" not in contexts:
            continue
        triggers = set(str(t).lower() for t in (fm.get("triggers") or []))
        applies = set(str(a).lower() for a in (fm.get("applies_to") or []))
        is_critical = fm.get("priority") == "critical"

        keep = False
        if tags & triggers:
            keep = True
        elif is_critical:
            keep = True
        elif applies & injury_areas and injury_areas:
            keep = True
        elif "any" in applies and tags:
            # Universal docs come along when the caller has at least one tag
            keep = True
        if keep:
            _maybe_add(doc_id, fm, body)

    # Sort selected docs by priority (critical first), then by id for determinism.
    PRIORITY_ORDER = {"critical": 0, "standard": 1, "reference": 2}
    selected.sort(key=lambda t: (PRIORITY_ORDER.get(t[1].get("priority", "standard"), 1), t[0]))

    # Budget-bounded combined text.
    combined_chunks: list[str] = []
    docs_pairs: list[tuple[str, str]] = []
    total = 0
    for doc_id, fm, body in selected:
        chunk = f"# {fm.get('title') or doc_id}\n\n{body}\n"
        if total + len(chunk) > max_chars and combined_chunks:
            # Already have at least one doc; stop adding.
            break
        combined_chunks.append(chunk)
        docs_pairs.append((doc_id, body))
        total += len(chunk)

    # Templates from block_library — filter by goal_tags overlap.
    templates: list[dict] = []
    try:
        from bot.services.db import list_block_library  # imported lazily for testability
        rows = list_block_library()
        for r in rows:
            gtags = set(str(g).lower() for g in (r.get("goal_tags") or []))
            if tags & gtags:
                templates.append(r)
    except Exception as e:
        logger.warning("program_gen: block_library fetch failed (degraded): %s", e)

    # Exemplars from data/knowledge/golden/
    exemplars = _load_golden_exemplars()

    kv = _knowledge_version(docs_pairs, templates, exemplars)

    payload = {
        "docs": docs_pairs,
        "templates": templates,
        "exemplars": exemplars,
        "knowledge_version": kv,
        "combined": "\n---\n".join(combined_chunks),
        "loaded_doc_ids": [d for d, _ in docs_pairs],
    }

    # Observability — log the call (same table the other surfaces use).
    if pitcher_id:
        _log_research_load(
            pitcher_id=pitcher_id,
            context="program_gen",
            trigger_reason=f"goal={','.join(sorted(tags)) or 'unspecified'}",
            doc_ids=payload["loaded_doc_ids"],
            total_chars=total,
            degraded=not templates and not exemplars,
        )

    return payload
